import openpyxl
import re
import string
import nltk,csv
import random
from nltk.corpus import stopwords
from nltk.stem.porter import PorterStemmer
from nltk.tokenize import RegexpTokenizer
from sklearn.model_selection import StratifiedKFold
from sklearn.linear_model import SGDClassifier
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer, TfidfTransformer
from sklearn import svm
from sklearn.metrics import accuracy_score,f1_score,precision_score,recall_score,classification_report,confusion_matrix
from sklearn.neural_network import MLPClassifier
from sklearn.naive_bayes import MultinomialNB
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import VotingClassifier
from sklearn.ensemble import RandomForestClassifier

import numpy as np

obamaTweetList = []
obamaClassLabelList = []
romneyTweetList = []
romneyClassLabelList = []
featureList = []
testTweetList = []
testLabelList = [] 
sheetName = ""

punc = string.punctuation.replace('_', '')
punc_regex = re.compile('[%s]' % re.escape(punc))
url_regex = '(http|ftp|https)://([\w_-]+(?:(?:\.[\w_-]+)+))([\w.,@?^=%&:/~+#-]*[\w@?^=%&/~+#-])?'
nonalpha1 = re.compile('[0-9]+[a-z]+')
nonalpha2 = re.compile('\s[0-9]+\s')
one_two = '\s+\w\w?\s+'


STOPWORDS = set(stopwords.words('english'))
STEMMER = PorterStemmer()
tokenizer = RegexpTokenizer(r'\w+')


def retrieveTweets(sheet):
    localLabelList = []
    localTweetList = []
    count = 0
    print("Sheet Name:",sheet)
	
    for row in range(3, sheet.max_row + 1):
        classLabel = str( sheet['E' + str(row)].value)
        tweet = sheet['D' + str(row)].value
        if(classLabel == "0" or classLabel == "1"or classLabel == "-1") and (tweet is not None ):
            localLabelList.append(int(classLabel))
            localTweetList.append(tweet)
            count = count + 1
    return localTweetList,localLabelList

def retrieveTestTweets(sheet):
    localLabelList = []
    localTweetList = []
    count = 0
    print("Sheet Name:",sheet)
	
    for row in range(3, sheet.max_row + 1):
        classLabel = str( sheet['E' + str(row)].value)
        tweet = sheet['A' + str(row)].value
        if(classLabel == "0" or classLabel == "1"or classLabel == "-1") and (tweet is not None ):
            localLabelList.append(int(classLabel))
            localTweetList.append(tweet)
            count = count + 1
    return localTweetList,localLabelList	
def removePunc(tweet):
    temp = re.sub(url_regex, ' ', tweet).strip()  # Remove URLs from the text
    temp = re.sub('@[\\w]*', ' ', temp).strip()
    temp = re.sub('rt',' ',temp).strip()
    temp = re.sub('yeah','yes',temp).strip()
    temp = re.sub('yess','yes',temp).strip()
    return tokenizer.tokenize(temp)

#start replaceTwoOrMore
def replaceTwoOrMore(s):
    #look for 2 or more repetitions of character and replace with the character itself
    pattern = re.compile(r"(.)\1{1,}", re.DOTALL)
    return pattern.sub(r"\1\1", s)
#end

def removealphanumeric(wordList):
    newWordList = []
    for word in wordList:
        newWord = re.sub('[^A-Za-z0-9]+', '', word)
        newWordList.append(newWord)
    return newWordList

def preProcess(tweetList):
    tempList = []
    for tweet in tweetList:
        tweet = tweet.lower()
        tweet = removePunc(tweet)
        tweet_removeNumbers = [item for item in tweet if not item.isdigit()]
        tweet_removeSingleCharacters = [s for s in tweet_removeNumbers if len(s) != 1]
#        tweet_removeObamaRomney = [s for s in tweet_removeSingleCharacters if s not in ('obama','romney')]
        tweet_removetwoormore = [replaceTwoOrMore(word) for word in tweet_removeSingleCharacters]
        tweet_alphanumberic = removealphanumeric(tweet_removetwoormore)
        tweet_nostopwords = [word for word in tweet_alphanumberic if not word in STOPWORDS]
        tweet_stemmed = [STEMMER.stem(word) for word in tweet_nostopwords]
        for word in tweet_stemmed:
            if(word == 'yess'):
                tweet_stemmed.remove(word)
                tweet_stemmed.append('yes')
        tempList.append(tweet_stemmed)
    return tempList
def AccessSheets():
    # OPEN THE EXCEL
    global sheetName
    wb = openpyxl.load_workbook('training-Obama-Romney-tweets.xlsx')
    obamaSheet = wb.get_sheet_by_name('Romney')
    global obamaTweetList, obamaClassLabelList
    obamaTweetList, obamaClassLabelList = retrieveTweets(obamaSheet)
    # SEND TWEETS FOR PREPROCESSING
    obamaTweetList = preProcess(obamaTweetList)

def saveProcessedTweets(tweetList,labelList):
    csvfile = "processed3.csv"
    # Assuming res is a flat list
    with open(csvfile, "w") as output:
        writer = csv.writer(output, lineterminator='\n')
        for val in tweetList:
            writer.writerow([val])
        for val in labelList:
            writer.writerow([val])	
	
def extract_features(tweet):
    global featureList
    tweet_words = set(tweet)
    features = {}
    for word in featureList:
        features['contains(%s)' % word] = (word in tweet_words)
    return features

def metricsComputation(predicted, xTestLabelList):
    Matrix = [[0 for x in range(3)] for y in range(3)]
    for tuple in range(0, len(predicted)):
        if (predicted[tuple] == xTestLabelList[tuple]):
            if (predicted[tuple] == 1):
                Matrix[0][0] = Matrix[0][0] + 1
            elif (predicted[tuple] == 0):
                Matrix[1][1] = Matrix[1][1] + 1
            elif (predicted[tuple] == -1):
                Matrix[2][2] = Matrix[2][2] + 1
        elif (predicted[tuple] == 1 and xTestLabelList[tuple] == 0):
            Matrix[1][0] = Matrix[1][0] + 1
        elif (predicted[tuple] == 1 and xTestLabelList[tuple] == -1):
            Matrix[2][0] = Matrix[2][0] + 1
        elif (predicted[tuple] == 0 and xTestLabelList[tuple] == 1):
            Matrix[0][1] = Matrix[0][1] + 1
        elif (predicted[tuple] == 0 and xTestLabelList[tuple] == -1):
            Matrix[2][1] = Matrix[2][1] + 1
        elif (predicted[tuple] == -1 and xTestLabelList[tuple] == 1):
            Matrix[0][2] = Matrix[0][2] + 1
        elif (predicted[tuple] == -1 and xTestLabelList[tuple] == 0):
            Matrix[1][2] = Matrix[1][2] + 1
    precisionPositive = (Matrix[0][0]) / (Matrix[0][0] + Matrix[1][0] + Matrix[2][0])
    recallPositive = (Matrix[0][0]) / (Matrix[0][0] + Matrix[0][1] + Matrix[0][2])
    precisionNegative = (Matrix[2][2]) / (Matrix[2][2] + Matrix[2][0] + Matrix[2][1])
    recallNegative = (Matrix[2][2]) / (Matrix[2][2] + Matrix[0][2] + Matrix[1][2])
    print("p+ " + str(precisionPositive))
    print("r+ " + str(recallPositive))
    print("r- " + str(recallNegative))
    print("p- " + str(precisionNegative))
    return precisionPositive,recallPositive,precisionNegative,recallNegative
def testModel():
    global sheetName
    wb = openpyxl.load_workbook('testing-Obama-Romney-tweets.xlsx')
    obamaSheet = wb.get_sheet_by_name('Romney')
    testTweetList, testLabelList = retrieveTestTweets(obamaSheet)
    # SEND TWEETS FOR PREPROCESSING
    testTweetList = preProcess(testTweetList)
    print(str(len(testTweetList)))
    return testTweetList,testLabelList
def TrainModel():
    global obamaTweetList, obamaClassLabelList
    ObamaFinalTweets = []
    xTestTweetList = []
	
    positivePercent = 0
    negativePercent = 0
    neutralPercent = 0
    positiveCnt = 0
    negativeCnt = 0
    neutralCnt = 0
	
    print("TRAINING DATA")
	#REMOVE TWEETS THAT DOESN NOT CONTAIN ANY WORDS
    for i in range(0, len(obamaTweetList)):
        if len(obamaTweetList[i]) <= 0:
            continue
        ObamaFinalTweets.append((obamaTweetList[i], obamaClassLabelList[i]))
	
    print("Before Scalling")			 
    print(str(len(ObamaFinalTweets)))
    print("Positive Tweets: ")
    posCnt = [x for x in ObamaFinalTweets if x[1] == 1]
    print(str(len(posCnt)*100/len(ObamaFinalTweets)))
    print("Negative Tweets: ")
    negCnt = [x for x in ObamaFinalTweets if x[1] == -1]
    print(str(len(negCnt)*100/len(ObamaFinalTweets)))
    print("Neutral Tweets: ")
    neutralCnt = [x for x in ObamaFinalTweets if x[1] == 0]
    print(str(len(neutralCnt)*100/len(ObamaFinalTweets)))  
	
    obamaScaledTweets = []
    positiveCnt = 0
    negativeCnt = 0
    neutralCnt = 0
	
    positivePercent = 35*len(ObamaFinalTweets)/100
    negativePercent = 35*len(ObamaFinalTweets)/100
    neutralPercent = 30*len(ObamaFinalTweets)/100
	
    for content in ObamaFinalTweets:
	     if(content[1] == 1 and positiveCnt < positivePercent):
		     obamaScaledTweets.append((content[0],content[1]))
		     positiveCnt = positiveCnt +1
	     elif(content[1] == -1 and negativeCnt < negativePercent):
		     obamaScaledTweets.append((content[0],content[1]))
		     negativeCnt = negativeCnt +1
	     elif(content[1] == 0 and neutralCnt < neutralPercent):
		     obamaScaledTweets.append((content[0],content[1]))
		     neutralCnt = neutralCnt + 1
    while(positiveCnt<positivePercent):
        posCnt = [x for x in ObamaFinalTweets if x[1] == 1]
        randomPositive = random.choice(posCnt)
        obamaScaledTweets.append((randomPositive[0],randomPositive[1]))
        positiveCnt = positiveCnt + 1
    while(neutralCnt<neutralPercent):
        neuCnt = [x for x in ObamaFinalTweets if x[1] == 1]
        randomNeutral = random.choice(neuCnt)
        obamaScaledTweets.append((randomNeutral[0],randomNeutral[1]))
        neutralCnt = neutralCnt + 1
    print("After Sampling: ")			 
    print(str(len(obamaScaledTweets)))
    print("Positive Tweets: ")
    posCnt = [x for x in obamaScaledTweets if x[1] == 1]
    print(str(len(posCnt)*100/len(obamaScaledTweets)))
    print("Negative Tweets: ")
    negCnt = [x for x in obamaScaledTweets if x[1] == -1]
    print(str(len(negCnt)*100/len(obamaScaledTweets)))
    print("Neutral Tweets: ")
    neutralCnt = [x for x in obamaScaledTweets if x[1] == 0]
    print(str(len(neutralCnt)*100/len(obamaScaledTweets)))     
		     
		     
    tweetList = []
    labelList = []
    for item in obamaScaledTweets:
        str1 = ' '.join(item[0])
        tweetList.append(str1.strip())
        labelList.append(item[1])
    count_vect = CountVectorizer(max_features = 1500)
	
#    saveProcessedTweets(tweetList, labelList)

	#CONVERTING THE TWEETS TO TF-IDF VECTOR FORM
    X = count_vect.fit_transform(tweetList)
    tfidf_transformer = TfidfTransformer()
    X_train_tfidf = tfidf_transformer.fit_transform(X)
	
	#TRAINING THE CLASSIFIER
    clf1 = SGDClassifier(alpha =0.001, learning_rate ='optimal', loss='epsilon_insensitive', penalty="l2",n_iter=100,random_state=42)
    clf2 = svm.LinearSVC(C=0.5,loss='hinge',random_state = 42)
    clf3 = RandomForestClassifier(n_estimators = 22,class_weight ='balanced_subsample',random_state=42,criterion ='gini')
	
    classifier = VotingClassifier(estimators=[('sgd', clf1), ('svm', clf2),('rc',clf3)], voting='hard')
    classifier.fit(X_train_tfidf, labelList)
	
	#TESTING TWEETS
    testTweetList, testLabelList = testModel()
    for item in testTweetList:
        str1 = ' '.join(item)
        xTestTweetList.append(str1.strip())
	
    X_new_counts = count_vect.transform(xTestTweetList)
    X_new_tfidf = tfidf_transformer.transform(X_new_counts)
    predicted = classifier.predict(X_new_tfidf)

    accuracy = accuracy_score(predicted,testLabelList)
    precisionPositive, recallPositive, precisionNegative, recallNegative = metricsComputation(predicted,testLabelList)
	
    print("Accuracy" + str(accuracy))
    print("POSITIVE CLASS: ")
    print("Precision: " + str(precisionPositive))
    print("Recall: " + str(recallPositive))
    print("F1: " + str((2 * precisionPositive * recallPositive) / (precisionPositive + recallPositive)))
    print("NEGATIVE CLASS: ")
    print("Precision: " + str(precisionNegative))
    print("Recall: " + str(recallNegative))
    print("F1: " + str((2 * precisionNegative * recallNegative) / (precisionNegative + recallNegative)))
    print("Tesing Done")
	
if __name__ == "__main__":
    AccessSheets()
    TrainModel()